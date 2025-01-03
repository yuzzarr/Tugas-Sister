from spyne import Application, rpc, ServiceBase, Integer, Unicode
from spyne.protocol.soap import Soap11
from spyne.server.wsgi import WsgiApplication
import openpyxl

class ExcelCRUDService(ServiceBase):
    # Create Operation: Adds a new record to the Excel file
    @rpc(Integer, Unicode, Unicode, Unicode, _returns=Unicode)
    def create(ctx, id, name, email, prodi):
        """Add a new record to the Excel file."""
        wb = openpyxl.load_workbook('data.xlsx')
        sheet = wb.active
        # Check for existing ID to prevent duplicates
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == id:
                return "ID already exists. Please use a different ID."
        sheet.append([id, name, email, prodi])
        wb.save('data.xlsx')
        return "Record created successfully"

    # Read Operation: Retrieves a record by ID
    @rpc(Integer, _returns=Unicode)
    def read(ctx, record_id):
        """Retrieve a record by ID from the Excel file."""
        wb = openpyxl.load_workbook('data.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == record_id:
                return f"ID: {row[0]}, Name: {row[1]}, Email: {row[2]}, Prodi: {row[3]}"
        return "Record not found"
    
     # Read All Operation: Retrieves all records from the Excel file
    @rpc(_returns=Unicode)
    def readAll(ctx):
        """Retrieve all records from the Excel file."""
        wb = openpyxl.load_workbook('data.xlsx')
        sheet = wb.active
        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            records.append(f"ID: {row[0]}, Name: {row[1]}, Email: {row[2]}, Prodi: {row[3]}")
        return "\n".join(records) if records else "No records found."

    # Update Operation: Updates a record by ID
    @rpc(Integer, Unicode, Unicode, Unicode, _returns=Unicode)
    def update(ctx, id, name, email, prodi):
        """Update a record by ID in the Excel file."""
        wb = openpyxl.load_workbook('data.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == id:
                row[1].value = name
                row[2].value = email
                row[3].value = prodi
                wb.save('data.xlsx')
                return "Record updated successfully"
        return "Record not found"

    # Delete Operation: Deletes a record by ID
    @rpc(Integer, _returns=Unicode)
    def delete(ctx, record_id):
        """Delete a record by ID from the Excel file."""
        wb = openpyxl.load_workbook('data.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == record_id:
                sheet.delete_rows(row[0].row)
                wb.save('data.xlsx')
                return "Record deleted successfully"
        return "Record not found"

# Setting up the SOAP application
application = Application([ExcelCRUDService],
    tns='spyne.examples.excel.soap',
    in_protocol=Soap11(validator='lxml'),
    out_protocol=Soap11()
)

if __name__ == '__main__':
    from wsgiref.simple_server import make_server
    server = make_server('127.0.0.1', 8000, WsgiApplication(application))
    print("SOAP server is running on http://127.0.0.1:8000")
    server.serve_forever()
